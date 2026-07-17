"""
jira_sprint_extract.py
======================
Extracts sprint issue data + full changelog from Jira Cloud and writes an Excel
workbook ready to paste into the Standup Assistant burndown chat.

Usage:
    python jira_sprint_extract.py --project ADA --sprint active
    python jira_sprint_extract.py --project ADA --sprint "260520-Sprint6"
    python jira_sprint_extract.py --project APO --sprint 2078       # sprint ID

Output:
    {PROJECT}_{sprint_slug}_extract_{YYYY-MM-DD}.xlsx

Sheets:
    Sprint Info     — sprint metadata (dates, goal, state)
    Issues          — one row per ticket, all key fields
    Transitions     — full status transition history (one row per transition)
    Rework          — Testing → In Progress transitions only (rework events)
    Scope Changes   — tickets added/removed mid-sprint from changelog

Setup:
    Set these environment variables (or edit the DEFAULTS block below):
        JIRA_BASE_URL   e.g. https://yoursite.atlassian.net
        JIRA_EMAIL      your Atlassian account email
        JIRA_API_TOKEN  API token from https://id.atlassian.com/manage-profile/security/api-tokens
"""

import os
import sys
import json
import argparse
import datetime
import re
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from dotenv import load_dotenv

load_dotenv()

# Console output contains Unicode (→ — · ✓). The default Windows console codepage
# (cp1252) can't encode these and raises UnicodeEncodeError, so force UTF-8 stdout.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass  # not a real text stream (e.g. captured/redirected) — nothing to do

# ── DEFAULTS (override via env vars or edit here) ────────────────────────────
DEFAULTS = {
    "base_url": os.environ.get("JIRA_BASE_URL", "https://sterlandsupport.atlassian.net"),
    "email":    os.environ.get("JIRA_EMAIL", ""),
    "token":    os.environ.get("JIRA_API_TOKEN", ""),
}

# ── Confluence sprint-commitment parent pages ─────────────────────────────────
CONFLUENCE_SPRINT_PARENTS = {
    "APO": "15172424",
    "ADA": "21987353",
    "ZEN": "25690233",
}

# ── Jira field mappings ───────────────────────────────────────────────────────
FIELD_SP        = "customfield_10059"
FIELD_DEV_HRS   = "customfield_10130"
FIELD_TEST_HRS  = "customfield_10131"
FIELD_SPRINT    = "customfield_10020"

# Status groupings (matching teams.json / system prompt)
# Burndown ladder (each line is the SP not yet past that milestone):
#   Dev Remaining    = BACKLOG ∪ DEV        (To Do, Team Estimated, BA Reviewed, In Progress, Blocked)
#   Review Remaining = Dev Remaining + Code Review   (not yet in Testing)
#   Remaining        = Review Remaining + TEST        (not yet Done)  == running_scope − done
# "Not Required" is excluded from every count.
BACKLOG = {"To Do", "BA Reviewed", "Team Estimated", "Open"}
DEV     = {"In Progress", "Blocked"}
REVIEW  = {"Code Review"}
TEST    = {"Testing", "Test Design"}
QA      = REVIEW | TEST          # all post-dev work (Code Review, Testing, Test Design)
DONE    = {"Done", "Closed"}

SYDNEY_OFFSET = datetime.timezone(datetime.timedelta(hours=10))


# ── API helpers ───────────────────────────────────────────────────────────────

def make_session(email, token):
    s = requests.Session()
    s.auth = HTTPBasicAuth(email, token)
    s.headers.update({"Accept": "application/json", "Content-Type": "application/json"})
    return s


def jira_get(session, base_url, path, params=None):
    if path.startswith("/agile/"):
        url = f"{base_url.rstrip('/')}/rest{path}"
    else:
        url = f"{base_url.rstrip('/')}/rest/api/3{path}"
    r = session.get(url, params=params)
    if r.status_code == 401:
        sys.exit("AUTH ERROR: Check your JIRA_EMAIL and JIRA_API_TOKEN.")
    if r.status_code == 400:
        sys.exit(f"BAD REQUEST: {r.text}")
    r.raise_for_status()
    return r.json()


def jira_post(session, base_url, path, body=None):
    url = f"{base_url.rstrip('/')}/rest/api/3{path}"
    r = session.post(url, json=body or {})
    if r.status_code == 401:
        sys.exit("AUTH ERROR: Check your JIRA_EMAIL and JIRA_API_TOKEN.")
    if r.status_code == 400:
        sys.exit(f"BAD REQUEST: {r.text}")
    r.raise_for_status()
    return r.json()


# ── Confluence helpers ────────────────────────────────────────────────────────

def confluence_get(session, base_url, path, params=None):
    url = f"{base_url.rstrip('/')}/wiki/rest/api{path}"
    r = session.get(url, params=params)
    if r.status_code == 401:
        sys.exit("AUTH ERROR: Check JIRA_EMAIL and JIRA_API_TOKEN.")
    if r.status_code in (403, 404):
        return None
    r.raise_for_status()
    return r.json()


def _parse_html_tables(html):
    """Extract tables from Confluence storage-format HTML → list[list[list[str]]]."""
    from html.parser import HTMLParser

    class _P(HTMLParser):
        def __init__(self):
            super().__init__()
            self.tables = []
            self._t = self._r = self._c = None
            self._in = False
            self._d = 0

        def handle_starttag(self, tag, attrs):
            t = tag.lower()
            if t == "table":
                if not self._d:
                    self._t = []
                    self.tables.append(self._t)
                self._d += 1
            elif t == "tr" and self._d == 1 and self._t is not None:
                self._r = []
                self._t.append(self._r)
            elif t in ("td", "th") and self._d == 1 and self._r is not None:
                self._c = []
                self._in = True
            elif t == "br" and self._in:
                self._c.append(" ")

        def handle_endtag(self, tag):
            t = tag.lower()
            if t == "table":
                self._d -= 1
                if not self._d:
                    self._t = None
            elif t in ("td", "th") and self._d == 1:
                if self._r is not None and self._c is not None:
                    self._r.append("".join(self._c).strip())
                self._in = False

        def handle_data(self, data):
            if self._in and self._c is not None:
                self._c.append(data.strip())

        def handle_entityref(self, name):
            if self._in and self._c is not None:
                ents = {"amp": "&", "lt": "<", "gt": ">", "nbsp": " ", "quot": '"', "apos": "'"}
                self._c.append(ents.get(name, ""))

        def handle_charref(self, name):
            if self._in and self._c is not None:
                try:
                    self._c.append(chr(int(name[1:], 16) if name.startswith("x") else int(name)))
                except Exception:
                    pass

    p = _P()
    p.feed(html)
    return p.tables


def _col_idx(headers, *keywords):
    """Return index of first header containing any keyword, else None."""
    for i, h in enumerate(headers):
        if any(k in h.lower() for k in keywords):
            return i
    return None


def parse_availability_from_confluence(html):
    """
    Find the team availability table in Confluence storage HTML.
    Returns list of member dicts or None if not found.
    """
    def safe_num(val):
        try:
            return float(str(val).replace(",", ".").strip())
        except (ValueError, AttributeError):
            return None

    for table in _parse_html_tables(html):
        if not table or len(table) < 2:
            continue
        headers = [c.lower() for c in table[0]]
        name_i    = _col_idx(headers, "name", "member", "person")
        planned_i = _col_idx(headers, "planned", "plan")
        if name_i is None or planned_i is None:
            continue
        actual_i   = _col_idx(headers, "actual")
        ncc_i      = _col_idx(headers, "ncc", "not counted", "exclude")
        full_cap_i = _col_idx(headers, "full capacity", "full cap", "full")

        members = []
        for row in table[1:]:
            name = row[name_i].strip() if len(row) > name_i else ""
            if not name or name.lower() in ("name", "total", "subtotal", ""):
                continue
            planned   = safe_num(row[planned_i])   if len(row) > planned_i else None
            actual    = safe_num(row[actual_i])    if actual_i    is not None and len(row) > actual_i    else None
            ncc_raw   = row[ncc_i].strip()         if ncc_i       is not None and len(row) > ncc_i       else ""
            full_cap  = safe_num(row[full_cap_i])  if full_cap_i  is not None and len(row) > full_cap_i  else None
            ncc = ncc_raw.lower() not in ("", "-", "no", "n", "0", "false")
            members.append({"name": name, "planned_days": planned, "actual_days": actual, "ncc": ncc,
                            "full_capacity": full_cap})

        if members:
            return members
    return None


def _cql_search_under(session, base_url, parent_id, cql_extra, expand=""):
    """Run a CQL search for descendants of parent_id with extra CQL clauses."""
    cql = f"ancestor = {parent_id} AND type = page AND {cql_extra}"
    params = {"cql": cql, "limit": 20}
    if expand:
        params["expand"] = expand
    return confluence_get(session, base_url, "/content/search", params=params)


def fetch_sprint_commitment(session, base_url, project, sprint_name, page_name=None):
    """
    Fetch and parse sprint commitment data from Confluence.
    Uses CQL ancestor search so Confluence folder-type nodes at intermediate
    levels don't block traversal (unlike /child/page which omits folders).
    Returns a commitment dict or None if unavailable.
    """
    parent_id = CONFLUENCE_SPRINT_PARENTS.get(project.upper())
    if not parent_id:
        return None

    matched = None

    if page_name:
        # Exact title first, then substring
        for clause in [f'title = "{page_name}"', f'title ~ "{page_name}"']:
            data = _cql_search_under(session, base_url, parent_id, clause,
                                     expand="body.storage,space")
            results = (data or {}).get("results", [])
            if results:
                matched = results[0]
                break
    else:
        # Auto-match: extract sprint number from sprint name (e.g. "260610-Sprint7" → 7)
        m = re.search(r"sprint\s*(\d+)", sprint_name, re.IGNORECASE)
        sprint_num = m.group(1) if m else None
        if sprint_num:
            clause = f'title ~ "Sprint {sprint_num}"'
            data = _cql_search_under(session, base_url, parent_id, clause,
                                     expand="body.storage,space")
            results = (data or {}).get("results", [])
            if results:
                # Prefer a title that contains the sprint number as a whole word
                for r in results:
                    if re.search(rf"\bsprint\s*{sprint_num}\b", r.get("title", ""), re.IGNORECASE):
                        matched = r
                        break
                if not matched:
                    matched = results[0]

    if not matched:
        label = page_name or sprint_name
        print(f"  No Confluence page found for '{label}'.")
        # Diagnostic: list available pages under parent
        diag = _cql_search_under(session, base_url, parent_id, "title ~ \"Sprint\"")
        titles = [r.get("title", "") for r in (diag or {}).get("results", [])]
        if titles:
            print(f"  Available pages (sample): {titles[:10]}")
        print(f"  Tip: pass --commitment-page \"<exact title>\" to specify it directly.")
        return None

    print(f"  Confluence page: {matched['title']}")
    html      = matched.get("body", {}).get("storage", {}).get("value", "")
    space_key = matched.get("space", {}).get("key", "")
    page_url  = f"{base_url}/wiki/spaces/{space_key}/pages/{matched['id']}"

    members = parse_availability_from_confluence(html)
    if not members:
        print("  Could not parse availability table from Confluence page.")
        return None

    # If the table has a Full Capacity column, Planned Days is already the 80% figure
    # so SP capacity = planned_excl_ncc * 2 (no extra × 0.8).
    # Without Full Capacity, the old formula (planned * 2 * 0.8) is preserved.
    has_full_capacity = any(m.get("full_capacity") is not None for m in members)
    planned_excl_ncc  = sum(m["planned_days"] or 0 for m in members if not m["ncc"])
    capacity_sp       = round(planned_excl_ncc * 2) if has_full_capacity else round((planned_excl_ncc * 2) * 0.8)
    planned_total     = sum(m["planned_days"] or 0 for m in members)
    has_actuals       = any(m["actual_days"] is not None for m in members)
    actual_total      = sum(m["actual_days"] or 0 for m in members if m["actual_days"] is not None)

    return {
        "page_title":         matched["title"],
        "page_url":           page_url,
        "members":            members,
        "planned_total":      planned_total,
        "planned_excl_ncc":   planned_excl_ncc,
        "capacity_sp":        capacity_sp,
        "has_actuals":        has_actuals,
        "actual_total":       actual_total,
        "has_full_capacity":  has_full_capacity,
    }


def resolve_sprint(session, base_url, project, sprint_arg):
    """
    Resolve sprint_arg to a sprint dict {id, name, state, startDate, endDate, goal}.
    sprint_arg can be: 'active', a sprint name substring, or a numeric sprint ID.
    """
    # Get all boards for the project
    boards_data = jira_get(session, base_url, f"/agile/1.0/board",
                           params={"projectKeyOrId": project, "maxResults": 50})
    boards = boards_data.get("values", [])
    if not boards:
        sys.exit(f"No boards found for project {project}.")

    board_id = boards[0]["id"]

    if sprint_arg.lower() == "active":
        data = jira_get(session, base_url, f"/agile/1.0/board/{board_id}/sprint",
                        params={"state": "active"})
        sprints = data.get("values", [])
        if not sprints:
            sys.exit(f"No active sprint found for project {project}.")
        return sprints[0]

    # Try numeric ID first
    if sprint_arg.isdigit():
        return jira_get(session, base_url, f"/agile/1.0/sprint/{sprint_arg}")

    # Fall back to name match — search recent sprints
    for state in ("active", "closed", "future"):
        data = jira_get(session, base_url, f"/agile/1.0/board/{board_id}/sprint",
                        params={"state": state, "maxResults": 50})
        for s in data.get("values", []):
            if sprint_arg.lower() in s["name"].lower():
                return s

    sys.exit(f"Sprint '{sprint_arg}' not found for project {project}.")


def fetch_issues_with_changelog(session, base_url, project, sprint_name):
    """
    Fetch all issues in the sprint with full changelog via expand=changelog.
    Returns list of issue dicts.
    """
    fields = [
        "summary", "status", "assignee", "issuetype", "priority",
        "resolution", "resolutiondate", "created", "updated",
        FIELD_SP, FIELD_DEV_HRS, FIELD_TEST_HRS, FIELD_SPRINT,
    ]
    all_issues = []
    next_page_token = None
    page_size = 50  # smaller page to stay within changelog payload limits

    jql = f'project = {project} AND sprint = "{sprint_name}"'

    while True:
        body = {
            "jql": jql,
            "fields": fields,
            "expand": "changelog",
            "maxResults": page_size,
        }
        if next_page_token:
            body["nextPageToken"] = next_page_token

        data = jira_post(session, base_url, "/search/jql", body=body)
        issues = data.get("issues", [])
        all_issues.extend(issues)
        total = data.get("total", "?")
        print(f"  Fetched {len(all_issues)} / {total} issues...")

        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break

    return all_issues


# ── Data extraction ───────────────────────────────────────────────────────────

def parse_dt(s):
    """Parse ISO-8601 string to aware datetime."""
    if not s:
        return None
    # Python < 3.11 doesn't handle all ISO variants — normalise Z
    s = s.replace("Z", "+00:00")
    return datetime.datetime.fromisoformat(s)


def to_sydney(dt):
    if dt is None:
        return None
    return dt.astimezone(SYDNEY_OFFSET)


def fmt_dt(dt):
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def status_phase(status_name):
    if status_name in BACKLOG:
        return "Backlog"
    if status_name in DEV:
        return "Dev"
    if status_name in QA:
        return "QA"
    if status_name in DONE:
        return "Done"
    return "Other"


def extract_transitions(issue):
    """
    Return list of dicts for every status transition in the changelog.
    """
    transitions = []
    key = issue["key"]
    changelog = issue.get("changelog", {})
    for entry in changelog.get("histories", []):
        ts = parse_dt(entry.get("created"))
        author = entry.get("author", {}).get("displayName", "")
        author_id = entry.get("author", {}).get("accountId", "")
        for item in entry.get("items", []):
            if item.get("field") == "status":
                transitions.append({
                    "key": key,
                    "timestamp": ts,
                    "timestamp_sydney": to_sydney(ts),
                    "author": author,
                    "author_id": author_id,
                    "from_status": item.get("fromString", ""),
                    "to_status": item.get("toString", ""),
                    "from_phase": status_phase(item.get("fromString", "")),
                    "to_phase": status_phase(item.get("toString", "")),
                })
    transitions.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))
    return transitions


def extract_sprint_membership_changes(issue, sprint_id):
    """
    Return list of dicts for sprint add/remove events from changelog.
    """
    changes = []
    key = issue["key"]
    changelog = issue.get("changelog", {})
    sprint_id_str = str(sprint_id)
    for entry in changelog.get("histories", []):
        ts = parse_dt(entry.get("created"))
        author = entry.get("author", {}).get("displayName", "")
        for item in entry.get("items", []):
            if item.get("field") == "Sprint":
                from_val = item.get("from", "") or ""
                to_val = item.get("to", "") or ""
                added   = sprint_id_str in to_val   and sprint_id_str not in from_val
                removed = sprint_id_str in from_val and sprint_id_str not in to_val
                if added or removed:
                    changes.append({
                        "key": key,
                        "timestamp": ts,
                        "timestamp_sydney": to_sydney(ts),
                        "author": author,
                        "change": "Added" if added else "Removed",
                        "from_raw": item.get("fromString", ""),
                        "to_raw": item.get("toString", ""),
                    })
    return changes


def extract_assignee_history(issue):
    """Return sorted list of {timestamp, to_assignee} from changelog assignee changes."""
    changes = []
    changelog = issue.get("changelog", {})
    for entry in changelog.get("histories", []):
        ts = parse_dt(entry.get("created"))
        for item in entry.get("items", []):
            if item.get("field") == "assignee":
                changes.append({
                    "timestamp": ts,
                    "to_assignee": item.get("toString") or "",
                })
    changes.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))
    return changes


def _assignee_at(initial, changes, ts):
    """Return the assignee name at or just before ts, given initial value and sorted change list."""
    current = initial
    for ch in changes:
        if ch["timestamp"] and ch["timestamp"] <= ts:
            current = ch["to_assignee"]
        else:
            break
    return current or "Unassigned"


def compute_allocations(issues, transitions_all, sprint):
    """
    Dev person  = assignee just before the ticket first entered QA or Done.
    Test person = assignee just before the ticket first entered Done,
                  or current assignee if still in Testing.

    "Not Required" tickets are excluded from all counts.

    Completed is based on CURRENT STATUS at extract time:
        dev  → current status is in QA (Code Review / Testing / Test Design) or Done
        test → current status is Done

    This gives the simple reconciliation the team expects:
        total_SP_excl_NR − Σ dev_completed_sp = dev_remaining_sp
        total_SP_excl_NR − Σ test_completed_sp = test_remaining_sp
    """
    initial_assignee = {
        iss["key"]: (iss["fields"].get("assignee") or {}).get("displayName", "") or "Unassigned"
        for iss in issues
    }
    issue_sp = {iss["key"]: iss["fields"].get(FIELD_SP) or 0 for iss in issues}
    current_status = {
        iss["key"]: iss["fields"].get("status", {}).get("name", "")
        for iss in issues
    }
    assignee_changes = {iss["key"]: extract_assignee_history(iss) for iss in issues}

    by_key = {}
    for t in transitions_all:
        by_key.setdefault(t["key"], []).append(t)
    for lst in by_key.values():
        lst.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))

    done_ts_map = {}
    for t in transitions_all:
        if t["to_status"] in DONE and t["key"] not in done_ts_map and t["timestamp"]:
            done_ts_map[t["key"]] = t["timestamp"]

    dev_alloc  = {}
    test_alloc = {}
    one_us = datetime.timedelta(microseconds=1)

    for iss in issues:
        key    = iss["key"]
        cur_st = current_status[key]

        # Not Required tickets are excluded from all allocation counts
        if cur_st == "Not Required":
            continue

        sp   = issue_sp[key]
        init = initial_assignee[key]
        ach  = assignee_changes.get(key, [])
        ts_list = by_key.get(key, [])

        # Current-status completion flags
        dev_done  = cur_st in QA | DONE   # currently in Code Review / Testing / Test Design or Done
        test_done = cur_st in DONE        # currently Done

        # First transition to QA or Done (dev hand-off point)
        dev_complete_ts = None
        for t in ts_list:
            if t["to_status"] in QA | DONE and t["timestamp"]:
                dev_complete_ts = t["timestamp"]
                break

        # First transition into QA (ticket entered testing)
        first_qa_ts = None
        for t in ts_list:
            if t["to_status"] in QA and t["timestamp"]:
                first_qa_ts = t["timestamp"]
                break

        done_ts = done_ts_map.get(key)

        # ── Dev allocation ─────────────────────────────────────────────────────
        # 1 µs before the hand-off captures the dev, not the incoming tester
        # (assignee change and status change share the same changelog timestamp).
        if dev_complete_ts:
            dev_person = _assignee_at(init, ach, dev_complete_ts - one_us)
        else:
            dev_person = _assignee_at(init, ach, datetime.datetime.now(datetime.timezone.utc))

        if dev_person not in dev_alloc:
            dev_alloc[dev_person] = {"keys": set(), "sp": 0, "completed_keys": set(), "completed_sp": 0}
        dev_alloc[dev_person]["keys"].add(key)
        dev_alloc[dev_person]["sp"] += sp
        if dev_done:
            dev_alloc[dev_person]["completed_keys"].add(key)
            dev_alloc[dev_person]["completed_sp"] += sp

        # ── Test allocation (only tickets that have entered QA) ────────────────
        if first_qa_ts:
            if done_ts:
                test_person = _assignee_at(init, ach, done_ts - one_us)
            else:
                test_person = _assignee_at(init, ach, datetime.datetime.now(datetime.timezone.utc))

            if test_person not in test_alloc:
                test_alloc[test_person] = {"keys": set(), "sp": 0, "completed_keys": set(), "completed_sp": 0}
            test_alloc[test_person]["keys"].add(key)
            test_alloc[test_person]["sp"] += sp
            if test_done:
                test_alloc[test_person]["completed_keys"].add(key)
                test_alloc[test_person]["completed_sp"] += sp

    return dev_alloc, test_alloc


def _business_hours_between(start_ts, end_ts, work_start_h=9, work_end_h=17, tz=SYDNEY_OFFSET):
    """Return working hours between two timestamps (Mon–Fri, 9 am–5 pm Sydney time)."""
    if not start_ts or not end_ts or end_ts <= start_ts:
        return 0.0
    start_local  = start_ts.astimezone(tz)
    end_local    = end_ts.astimezone(tz)
    total        = 0.0
    current_date = start_local.date()
    end_date     = end_local.date()
    while current_date <= end_date:
        if current_date.weekday() < 5:  # Monday–Friday
            day_start = datetime.datetime.combine(current_date, datetime.time(work_start_h), tzinfo=tz)
            day_end   = datetime.datetime.combine(current_date, datetime.time(work_end_h),   tzinfo=tz)
            seg_start = max(start_local, day_start)
            seg_end   = min(end_local,   day_end)
            if seg_end > seg_start:
                total += (seg_end - seg_start).total_seconds() / 3600
        current_date += datetime.timedelta(days=1)
    return total


def _nearest_fibonacci(value):
    """Snap a positive float to the nearest Fibonacci number (1, 2, 3, 5, 8, 13, 21, …)."""
    if value <= 0:
        return 0
    fibs = [1, 2, 3, 5, 8, 13, 21, 34, 55, 89]
    while fibs[-1] < value:
        fibs.append(fibs[-1] + fibs[-2])
    return min(fibs, key=lambda f: abs(f - value))


def compute_cycle_times(issues, transitions_all, sprint):
    """
    Compute total hours each issue spent in key statuses across its full changelog.
    Time in the same status across multiple visits is accumulated (e.g. two "In Progress"
    stints of 8h and 20h → 28h total).

    Tracked buckets:
        In Progress  → "In Progress"
        Blocked      → "Blocked"
        Code Review  → "Code Review"
        Test         → "Testing" + "Test Design" combined
    """
    STATUS_BUCKET = {
        "in progress": "In Progress",
        "blocked":     "Blocked",
        "code review": "Code Review",
        "testing":     "Test",
        "test design": "Test",
    }

    now_utc    = datetime.datetime.now(datetime.timezone.utc)
    sprint_end = parse_dt(sprint.get("endDate"))
    effective_now = sprint_end if (sprint_end and sprint_end < now_utc) else now_utc

    by_key = {}
    for t in transitions_all:
        by_key.setdefault(t["key"], []).append(t)
    for lst in by_key.values():
        lst.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))

    result = {}
    for issue in issues:
        key    = issue["key"]
        totals = {"In Progress": 0.0, "Blocked": 0.0, "Code Review": 0.0, "Test": 0.0}
        trans  = by_key.get(key, [])
        if not trans:
            result[key] = totals
            continue

        # Build (status, start_ts, end_ts) segments from the changelog.
        segments = []

        # Initial segment: issue creation → first recorded transition, in the from_status.
        # Only include it when the initial status is UNTRACKED (e.g. "To Do").
        # If the first from_status is a tracked status (e.g. "In Progress"), the Jira search
        # API likely returned a truncated changelog and using the creation date as the start
        # would silently absorb all pre-sprint backlog time into that tracked bucket.
        created_ts     = parse_dt(issue["fields"].get("created"))
        first_ts       = trans[0]["timestamp"]
        initial_status = trans[0]["from_status"]
        if (created_ts and first_ts and first_ts > created_ts
                and not STATUS_BUCKET.get(initial_status.lower())):
            segments.append((initial_status, created_ts, first_ts))

        # Inter-transition segments
        for i in range(len(trans) - 1):
            t1, t2 = trans[i], trans[i + 1]
            if t1["timestamp"] and t2["timestamp"]:
                segments.append((t1["to_status"], t1["timestamp"], t2["timestamp"]))

        # Open-ended final segment up to sprint end / now (only when not Done/Closed)
        last_t = trans[-1]
        if last_t["timestamp"] and last_t["to_status"] not in DONE:
            segments.append((last_t["to_status"], last_t["timestamp"], effective_now))

        for status, start, end in segments:
            bucket = STATUS_BUCKET.get(status.lower())
            if bucket and start and end:
                totals[bucket] += _business_hours_between(start, end)

        result[key] = totals

    return result


# ── Excel builder ─────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", start_color="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT  = Font(name="Arial", size=10)
ALT_FILL   = PatternFill("solid", start_color="EBF3FB")
REWORK_FILL = PatternFill("solid", start_color="FCE4D6")
WARN_FILL  = PatternFill("solid", start_color="FFF2CC")
THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row=1, max_col=None):
    max_col = max_col or ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_data_row(ws, row, max_col, alt=False, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if fill:
            cell.fill = fill
        elif alt:
            cell.fill = ALT_FILL


def autofit(ws, min_width=8, max_width=50):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def add_commitment_sheet(wb, sprint, commitment):
    """Add a Commitment sheet with team availability and capacity calculation."""
    ws = wb.create_sheet("Commitment")
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Sprint Commitment — {sprint.get('name', '')}"
    c.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    c.fill  = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    if not commitment:
        ws["A2"] = "No commitment data found in Confluence."
        ws["A2"].font = Font(italic=True, color="999999", name="Arial")
        return

    # Sprint dates (row 2)
    sprint_start = parse_dt(sprint.get("startDate"))
    sprint_end   = parse_dt(sprint.get("endDate"))
    start_str = sprint_start.astimezone(SYDNEY_OFFSET).strftime("%d %b %Y") if sprint_start else ""
    end_str   = sprint_end.astimezone(SYDNEY_OFFSET).strftime("%d %b %Y")   if sprint_end   else ""

    ws["A2"] = "Sprint Period:"
    ws["A2"].font = Font(bold=True, name="Arial", size=10)
    ws["B2"] = f"{start_str}  –  {end_str}"
    ws["B2"].font = Font(name="Arial", size=10)
    ws["D2"] = "Source:"
    ws["D2"].font = Font(bold=True, name="Arial", size=10)
    ws["E2"] = commitment["page_title"]
    ws["E2"].font = Font(name="Arial", size=10)
    ws.row_dimensions[2].height = 16

    # Column headers (row 4)
    has_actuals   = commitment.get("has_actuals", False)
    has_full_cap  = commitment.get("has_full_capacity", False)
    col_map = {"name": 1}
    next_col = 2
    if has_full_cap:
        col_map["full_cap"] = next_col;  next_col += 1
    col_map["planned"] = next_col;       next_col += 1
    if has_actuals:
        col_map["actual"] = next_col;    next_col += 1
    col_map["ncc"]     = next_col;       next_col += 1
    col_map["counted"] = next_col;       next_col += 1
    col_map["sp"]      = next_col

    headers_row = {
        col_map["name"]:    "Team Member",
        col_map["planned"]: "Planned Days",
        col_map["ncc"]:     "NCC",
        col_map["counted"]: "Counted Days",
        col_map["sp"]:      "SP Contribution",
    }
    if has_actuals:
        headers_row[col_map["actual"]] = "Actual Days"
    if has_full_cap:
        headers_row[col_map["full_cap"]] = "Full Capacity"

    n_cols = col_map["sp"]
    for col_i, label in headers_row.items():
        c = ws.cell(4, col_i, label)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[4].height = 24

    first_data_row = 5
    last_data_row  = 4 + len(commitment["members"])

    # Member rows
    for row_i, member in enumerate(commitment["members"], first_data_row):
        ncc     = member["ncc"]
        planned = member["planned_days"]

        b_col = get_column_letter(col_map["planned"])
        d_col = get_column_letter(col_map["ncc"])
        e_col = get_column_letter(col_map["counted"])

        ws.cell(row_i, col_map["name"],    member["name"])
        if has_full_cap:
            ws.cell(row_i, col_map["full_cap"], member.get("full_capacity"))
        ws.cell(row_i, col_map["planned"], planned)
        # NCC: boolean TRUE/FALSE so the IF formula can reference it directly
        ws.cell(row_i, col_map["ncc"]).value = True if ncc else False
        # Counted Days: =IF(D_row=TRUE, 0, B_row)
        ws.cell(row_i, col_map["counted"]).value = (
            f"=IF({d_col}{row_i}=TRUE,0,{b_col}{row_i})"
        )
        # SP Contribution: planned is already 80% when full_capacity column is present
        if has_full_cap:
            ws.cell(row_i, col_map["sp"]).value = f"={e_col}{row_i}*2"
        else:
            ws.cell(row_i, col_map["sp"]).value = f"=({e_col}{row_i}*2)*0.8"
        if has_actuals:
            ws.cell(row_i, col_map["actual"], member.get("actual_days"))

        for col_i in range(1, n_cols + 1):
            c = ws.cell(row_i, col_i)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center" if col_i > 1 else "left", vertical="center")
        alt = (row_i % 2 == 0)
        style_data_row(ws, row_i, n_cols, alt=alt)

    # Totals row
    total_row = last_data_row + 1
    b_col = get_column_letter(col_map["planned"])
    e_col = get_column_letter(col_map["counted"])
    f_col = get_column_letter(col_map["sp"])

    ws.cell(total_row, col_map["name"]).value = "TOTAL"
    if has_full_cap:
        fc_col = get_column_letter(col_map["full_cap"])
        ws.cell(total_row, col_map["full_cap"]).value = (
            f"=SUM({fc_col}{first_data_row}:{fc_col}{last_data_row})"
        )
    ws.cell(total_row, col_map["planned"]).value = (
        f"=SUM({b_col}{first_data_row}:{b_col}{last_data_row})"
    )
    ws.cell(total_row, col_map["counted"]).value = (
        f"=SUM({e_col}{first_data_row}:{e_col}{last_data_row})"
    )
    ws.cell(total_row, col_map["sp"]).value = (
        f"=SUM({f_col}{first_data_row}:{f_col}{last_data_row})"
    )
    if has_actuals:
        c_col = get_column_letter(col_map["actual"])
        ws.cell(total_row, col_map["actual"]).value = (
            f"=SUM({c_col}{first_data_row}:{c_col}{last_data_row})"
        )
    for col_i in range(1, n_cols + 1):
        c = ws.cell(total_row, col_i)
        c.font   = Font(bold=True, name="Arial", size=10)
        c.fill   = PatternFill("solid", start_color="D9E1F2")
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center" if col_i > 1 else "left", vertical="center")

    # Capacity formula note
    cap_row = total_row + 2
    ws.merge_cells(f"A{cap_row}:G{cap_row}")
    f_col = get_column_letter(col_map["sp"])
    if has_full_cap:
        formula = f"=\"Capacity = \"&{e_col}{total_row}&\" counted days × 2 SP/day  =  \"&{f_col}{total_row}&\" SP\""
    else:
        formula = f"=\"Capacity = (\"&{e_col}{total_row}&\" counted days × 2 SP/day) × 80%  =  \"&{f_col}{total_row}&\" SP\""
    c = ws.cell(cap_row, 1)
    c.value = formula
    c.font = Font(name="Arial", size=10, italic=True, color="4D504E")

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col_i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14


def add_allocation_sheet(wb, dev_alloc, test_alloc):
    """Add an Allocation sheet with developer and tester story-point summaries."""
    ws = wb.create_sheet("Allocation")

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Sprint Allocation"
    c.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    c.fill  = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    COL_HEADERS = ["Assignee", "Tickets", "Story Points", "Completed Tickets", "Completed SP"]
    N = len(COL_HEADERS)
    BLUE  = PatternFill("solid", start_color="006CA7")
    GREEN = PatternFill("solid", start_color="507E35")

    def _write_section(start_row, label, alloc, section_fill):
        r = start_row
        ws.merge_cells(f"A{r}:E{r}")
        sec = ws.cell(r, 1, label)
        sec.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        sec.fill = section_fill
        sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        r += 1

        for ci, h in enumerate(COL_HEADERS, 1):
            hc = ws.cell(r, ci, h)
            hc.fill = HDR_FILL
            hc.font = HDR_FONT
            hc.border = THIN_BORDER
            hc.alignment = Alignment(
                horizontal="center" if ci > 1 else "left", vertical="center"
            )
        ws.row_dimensions[r].height = 20
        r += 1

        if not alloc:
            ws.cell(r, 1, "No data").font = Font(italic=True, color="999999", name="Arial")
            return r + 1

        rows = sorted(alloc.items(), key=lambda x: (-x[1]["completed_sp"], x[0]))
        for i, (name, st) in enumerate(rows):
            ws.cell(r, 1).value = name
            ws.cell(r, 2).value = len(st["keys"])
            ws.cell(r, 3).value = st["sp"]
            ws.cell(r, 4).value = len(st["completed_keys"])
            ws.cell(r, 5).value = st["completed_sp"]
            style_data_row(ws, r, N, alt=(i % 2 == 1))
            ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
            for ci in range(2, N + 1):
                ws.cell(r, ci).alignment = Alignment(horizontal="center", vertical="center")
            r += 1
        return r

    row = 2
    row = _write_section(row, "Developer Allocation  ·  Completed = currently in Testing or Done", dev_alloc, BLUE)
    row += 1
    _write_section(row, "Tester Allocation  ·  Completed = currently Done", test_alloc, GREEN)

    ws.column_dimensions["A"].width = 28
    for ci in range(2, N + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def build_workbook(sprint, issues, transitions_all, rework_all, scope_changes_all, commitment=None):
    wb = Workbook()

    # ── Sheet 1: Sprint Info ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Sprint Info"
    ws.freeze_panes = "B2"

    info_rows = [
        ("Sprint Name",  sprint.get("name", "")),
        ("Sprint ID",    sprint.get("id", "")),
        ("State",        sprint.get("state", "")),
        ("Start Date",   fmt_dt(to_sydney(parse_dt(sprint.get("startDate"))))),
        ("End Date",     fmt_dt(to_sydney(parse_dt(sprint.get("endDate"))))),
        ("Goal",         sprint.get("goal", "")),
        ("Total Issues", len(issues)),
    ]
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    for r, (k, v) in enumerate(info_rows, start=1):
        ws.cell(r, 1, k).font = Font(bold=True, name="Arial", size=10)
        ws.cell(r, 2, v).font = BODY_FONT

    # ── Sheet 2: Issues ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues")
    ws2.freeze_panes = "A2"

    issue_headers = [
        "Key", "Summary", "Type", "Status", "Phase", "Priority",
        "SP", "Dev Hours Est", "Test Hours Est",
        "Assignee", "Reporter",
        "Created (AEST)", "Updated (AEST)", "Resolution Date (AEST)",
        "Resolution", "Committed (Y/N)", "Sprint Added Date (AEST)",
    ]
    ws2.append(issue_headers)
    style_header_row(ws2, max_col=len(issue_headers))

    sprint_start = parse_dt(sprint.get("startDate"))
    grace_cutoff = sprint_start + datetime.timedelta(hours=12) if sprint_start else None

    for i, issue in enumerate(issues):
        f = issue["fields"]
        created = parse_dt(f.get("created"))
        committed = "Y" if (grace_cutoff and created and created <= grace_cutoff) else "N"

        # Find first sprint-add event for this issue
        sprint_add_dt = None
        for sc in scope_changes_all:
            if sc["key"] == issue["key"] and sc["change"] == "Added":
                sprint_add_dt = sc["timestamp_sydney"]
                break

        row = [
            issue["key"],
            f.get("summary", ""),
            f.get("issuetype", {}).get("name", ""),
            f.get("status", {}).get("name", ""),
            status_phase(f.get("status", {}).get("name", "")),
            f.get("priority", {}).get("name", ""),
            f.get(FIELD_SP),
            f.get(FIELD_DEV_HRS),
            f.get(FIELD_TEST_HRS),
            (f.get("assignee") or {}).get("displayName", ""),
            (f.get("reporter") or {}).get("displayName", ""),
            fmt_dt(to_sydney(created)),
            fmt_dt(to_sydney(parse_dt(f.get("updated")))),
            fmt_dt(to_sydney(parse_dt(f.get("resolutiondate")))),
            (f.get("resolution") or {}).get("name", ""),
            committed,
            fmt_dt(sprint_add_dt) if sprint_add_dt else "",
        ]
        ws2.append(row)
        alt = (i % 2 == 1)
        style_data_row(ws2, i + 2, len(issue_headers), alt=alt)

    autofit(ws2)
    ws2.column_dimensions["B"].width = 52  # summary

    # ── Sheet 3: Transitions ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Transitions")
    ws3.freeze_panes = "A2"

    trans_headers = [
        "Key", "Timestamp (AEST)", "Author",
        "From Status", "From Phase", "To Status", "To Phase",
        "Is Rework (Testing→not Done)",
    ]
    ws3.append(trans_headers)
    style_header_row(ws3, max_col=len(trans_headers))

    for i, t in enumerate(transitions_all):
        is_rework = (t["from_status"] == "Testing" and t["to_status"] != "Done")
        row = [
            t["key"],
            fmt_dt(t["timestamp_sydney"]),
            t["author"],
            t["from_status"],
            t["from_phase"],
            t["to_status"],
            t["to_phase"],
            "YES" if is_rework else "",
        ]
        ws3.append(row)
        fill = REWORK_FILL if is_rework else None
        style_data_row(ws3, i + 2, len(trans_headers), alt=(i % 2 == 1), fill=fill)

    autofit(ws3)

    # ── Sheet 4: Rework ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Rework")
    ws4.freeze_panes = "A2"

    rework_headers = [
        "Key", "Summary", "SP", "Assignee",
        "Rework Timestamp (AEST)", "Rework #", "Author (who sent back)",
    ]
    ws4.append(rework_headers)
    style_header_row(ws4, max_col=len(rework_headers))

    # Count rework per ticket
    rework_counts = {}
    for r in rework_all:
        rework_counts[r["key"]] = rework_counts.get(r["key"], 0) + 1

    ticket_rework_idx = {}
    for i, r in enumerate(rework_all):
        key = r["key"]
        ticket_rework_idx[key] = ticket_rework_idx.get(key, 0) + 1
        issue_fields = next((iss["fields"] for iss in issues if iss["key"] == key), {})
        row = [
            key,
            issue_fields.get("summary", ""),
            issue_fields.get(FIELD_SP),
            (issue_fields.get("assignee") or {}).get("displayName", ""),
            fmt_dt(r["timestamp_sydney"]),
            ticket_rework_idx[key],
            r["author"],
        ]
        ws4.append(row)
        style_data_row(ws4, i + 2, len(rework_headers), fill=REWORK_FILL)

    if not rework_all:
        ws4.append(["No rework events detected in this sprint."])

    autofit(ws4)
    ws4.column_dimensions["B"].width = 50

    # ── Sheet 5: Scope Changes ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Scope Changes")
    ws5.freeze_panes = "A2"

    scope_headers = [
        "Key", "Summary", "SP", "Change", "Timestamp (AEST)", "Author",
    ]
    ws5.append(scope_headers)
    style_header_row(ws5, max_col=len(scope_headers))

    # Combine sprint membership changes with Not Required transitions (scope removals)
    nr_rows_ws5 = [
        {"key": t["key"], "change": "Not Required",
         "timestamp_sydney": t["timestamp_sydney"], "author": t["author"]}
        for t in transitions_all if t["to_status"] == "Not Required"
    ]
    all_scope_rows_ws5 = sorted(
        scope_changes_all + nr_rows_ws5,
        key=lambda x: x.get("timestamp_sydney") or datetime.datetime.min.replace(tzinfo=SYDNEY_OFFSET)
    )

    NR_FILL = PatternFill("solid", start_color="F2DCDB")
    for i, sc in enumerate(all_scope_rows_ws5):
        issue_f = next((iss["fields"] for iss in issues if iss["key"] == sc["key"]), {})
        row = [
            sc["key"],
            issue_f.get("summary", ""),
            issue_f.get(FIELD_SP),
            sc["change"],
            fmt_dt(sc.get("timestamp_sydney")),
            sc.get("author", ""),
        ]
        ws5.append(row)
        if sc["change"] == "Added":
            fill = WARN_FILL
        elif sc["change"] == "Not Required":
            fill = NR_FILL
        else:
            fill = None
        style_data_row(ws5, i + 2, len(scope_headers), alt=(i % 2 == 1), fill=fill)

    if not all_scope_rows_ws5:
        ws5.append(["No mid-sprint scope changes detected."])

    autofit(ws5)
    ws5.column_dimensions["B"].width = 50

    # ── Commitment ────────────────────────────────────────────────────────────
    if commitment:
        add_commitment_sheet(wb, sprint, commitment)

    # ── Allocation ────────────────────────────────────────────────────────────
    dev_alloc, test_alloc = compute_allocations(issues, transitions_all, sprint)
    add_allocation_sheet(wb, dev_alloc, test_alloc)

    # ── Dashboard + charts ────────────────────────────────────────────────────
    cs          = compute_chart_series(sprint, issues, transitions_all, rework_all, scope_changes_all)
    cycle_times = compute_cycle_times(issues, transitions_all, sprint)
    add_dashboard_sheet(wb, sprint, cs, commitment,
                        scope_changes_all=scope_changes_all,
                        transitions_all=transitions_all,
                        rework_all=rework_all,
                        cycle_times=cycle_times)

    return wb


# ── Paste-ready summary for the chat ─────────────────────────────────────────

def build_chat_summary(sprint, issues, transitions_all, rework_all, scope_changes_all, commitment=None):
    """
    Returns a markdown string the SM can paste into the burndown chat.
    Structured so the assistant can reconstruct all four series without
    needing to re-query Jira.
    """
    sprint_start = parse_dt(sprint.get("startDate"))
    sprint_end   = parse_dt(sprint.get("endDate"))
    grace_cutoff = sprint_start + datetime.timedelta(hours=12) if sprint_start else None

    lines = []
    lines.append(f"## Sprint Data: {sprint['name']}")
    lines.append(f"**Sprint ID:** {sprint['id']}  ")
    lines.append(f"**State:** {sprint['state']}  ")
    lines.append(f"**Start:** {fmt_dt(to_sydney(sprint_start))} AEST  ")
    lines.append(f"**End:** {fmt_dt(to_sydney(sprint_end))} AEST  ")
    lines.append(f"**Goal:** {sprint.get('goal', '(none)')}  ")
    lines.append("")

    lines.append("### Issues")
    lines.append("| Key | Summary | Type | Status | SP | Dev Hrs | Test Hrs | Assignee | Committed | Resolution Date (AEST) |")
    lines.append("|-----|---------|------|--------|----|---------|----------|----------|-----------|----------------------|")
    for iss in issues:
        f = iss["fields"]
        created = parse_dt(f.get("created"))
        committed = "Y" if (grace_cutoff and created and created <= grace_cutoff) else "N"
        lines.append(
            f"| {iss['key']} | {f.get('summary','')[:60]} | "
            f"{f.get('issuetype',{}).get('name','')} | "
            f"{f.get('status',{}).get('name','')} | "
            f"{f.get(FIELD_SP) or 0} | "
            f"{f.get(FIELD_DEV_HRS) or 0} | "
            f"{f.get(FIELD_TEST_HRS) or 0} | "
            f"{(f.get('assignee') or {}).get('displayName','')} | "
            f"{committed} | "
            f"{fmt_dt(to_sydney(parse_dt(f.get('resolutiondate'))))} |"
        )
    lines.append("")

    lines.append("### Status Transitions (full history)")
    lines.append("| Key | Timestamp (AEST) | From | To | Is Rework |")
    lines.append("|-----|-----------------|------|----|-----------|")
    for t in transitions_all:
        is_rework = t["from_status"] == "Testing" and t["to_status"] != "Done"
        lines.append(
            f"| {t['key']} | {fmt_dt(t['timestamp_sydney'])} | "
            f"{t['from_status']} | {t['to_status']} | "
            f"{'YES' if is_rework else ''} |"
        )
    lines.append("")

    lines.append("### Rework Events (Testing → In Progress only)")
    if rework_all:
        lines.append("| Key | Timestamp (AEST) | Author |")
        lines.append("|-----|-----------------|--------|")
        for r in rework_all:
            lines.append(f"| {r['key']} | {fmt_dt(r['timestamp_sydney'])} | {r['author']} |")
    else:
        lines.append("_No rework events detected._")
    lines.append("")

    lines.append("### Scope Changes (mid-sprint adds/removes/not-required)")

    # Gather Not Required transitions as scope removal events
    nr_scope_rows = []
    for t in transitions_all:
        if t["to_status"] == "Not Required":
            nr_scope_rows.append({
                "key": t["key"], "change": "Not Required",
                "timestamp_sydney": t["timestamp_sydney"], "author": t["author"],
            })

    all_scope_rows = sorted(
        scope_changes_all + nr_scope_rows,
        key=lambda x: x.get("timestamp_sydney") or datetime.datetime.min.replace(tzinfo=SYDNEY_OFFSET)
    )
    if all_scope_rows:
        lines.append("| Key | Change | Timestamp (AEST) | Author |")
        lines.append("|-----|--------|-----------------|--------|")
        for sc in all_scope_rows:
            lines.append(f"| {sc['key']} | {sc['change']} | {fmt_dt(sc.get('timestamp_sydney'))} | {sc.get('author','')} |")
    else:
        lines.append("_No scope changes detected._")
    lines.append("")

    # ── Commitment (Confluence) ───────────────────────────────────────────────
    if commitment:
        lines.append("### Sprint Commitment (Confluence)")
        lines.append(f"**Source:** {commitment['page_title']}  ")
        has_full_cap = commitment.get("has_full_capacity", False)
        if has_full_cap:
            lines.append(f"**Capacity:** {commitment['planned_excl_ncc']} counted days × 2 SP/day = **{commitment['capacity_sp']} SP**  ")
        else:
            lines.append(f"**Capacity:** ({commitment['planned_excl_ncc']} counted days × 2 SP/day) × 80% = **{commitment['capacity_sp']} SP**  ")
        lines.append("")
        has_actuals = commitment.get("has_actuals", False)
        if has_full_cap and has_actuals:
            lines.append("| Team Member | Full Capacity | Planned Days | Actual Days | NCC | Counted Days | SP Contrib |")
            lines.append("|-------------|--------------|-------------|------------|-----|-------------|-----------|")
        elif has_full_cap:
            lines.append("| Team Member | Full Capacity | Planned Days | NCC | Counted Days | SP Contrib |")
            lines.append("|-------------|--------------|-------------|-----|-------------|-----------|")
        elif has_actuals:
            lines.append("| Team Member | Planned Days | Actual Days | NCC | Counted Days | SP Contrib |")
            lines.append("|-------------|-------------|------------|-----|-------------|-----------|")
        else:
            lines.append("| Team Member | Planned Days | NCC | Counted Days | SP Contrib |")
            lines.append("|-------------|-------------|-----|-------------|-----------|")
        for m in commitment["members"]:
            counted  = 0 if m["ncc"] else (m["planned_days"] or 0)
            sp_c     = round(counted * 2, 1) if has_full_cap else round(counted * 2 * 0.8, 1)
            ncc_str  = "✓" if m["ncc"] else ""
            fc_str   = str(m.get("full_capacity") or "") if has_full_cap else None
            if has_full_cap and has_actuals:
                lines.append(
                    f"| {m['name']} | {fc_str} | {m['planned_days'] or ''} | {m['actual_days'] or ''} "
                    f"| {ncc_str} | {counted} | {sp_c} |"
                )
            elif has_full_cap:
                lines.append(
                    f"| {m['name']} | {fc_str} | {m['planned_days'] or ''} | {ncc_str} | {counted} | {sp_c} |"
                )
            elif has_actuals:
                lines.append(
                    f"| {m['name']} | {m['planned_days'] or ''} | {m['actual_days'] or ''} "
                    f"| {ncc_str} | {counted} | {sp_c} |"
                )
            else:
                lines.append(
                    f"| {m['name']} | {m['planned_days'] or ''} | {ncc_str} | {counted} | {sp_c} |"
                )
        fc_total_str = f"**{sum(m.get('full_capacity') or 0 for m in commitment['members'])}** | " if has_full_cap else ""
        lines.append(f"| **TOTAL** | {fc_total_str}**{commitment['planned_total']}** | "
                     + ("**—** | " if has_actuals else "")
                     + f"| **{commitment['planned_excl_ncc']}** | **{commitment['capacity_sp']}** |")

    return "\n".join(lines)


# ── Dashboard charts ──────────────────────────────────────────────────────────

def _get_business_days(start_date, end_date):
    """Return Mon–Fri dates from start_date to end_date inclusive."""
    days, d = [], start_date
    while d <= end_date:
        if d.weekday() < 5:
            days.append(d)
        d += datetime.timedelta(days=1)
    return days


def compute_chart_series(sprint, issues, transitions_all, rework_all, scope_changes_all):
    """Compute burndown, scope-change and rework series from extracted sprint data.

    Burndown scope is dynamic: starts at committed (grace-window) SP, then rises
    when tickets are added and falls when tickets are removed or set Not Required.
    Done detection uses changelog transitions only — no resolutiondate gate.
    """
    sprint_start = parse_dt(sprint.get("startDate"))
    sprint_end   = parse_dt(sprint.get("endDate"))
    if not sprint_start or not sprint_end:
        return None

    today_utc     = datetime.datetime.now(datetime.timezone.utc)
    today_syd     = today_utc.astimezone(SYDNEY_OFFSET).date()
    start_date    = sprint_start.astimezone(SYDNEY_OFFSET).date()
    end_date      = sprint_end.astimezone(SYDNEY_OFFSET).date()
    effective_end = min(end_date, today_syd)

    biz_days = _get_business_days(start_date, end_date)
    n = len(biz_days)
    if n == 0:
        return None

    # ── Build lookups ─────────────────────────────────────────────────────────
    issue_fields = {iss["key"]: iss["fields"] for iss in issues}
    issue_sp     = {key: (f.get(FIELD_SP) or 0) for key, f in issue_fields.items()}

    by_key = {}
    for t in transitions_all:
        by_key.setdefault(t["key"], []).append(t)
    for lst in by_key.values():
        lst.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))

    current_status = {k: f.get("status", {}).get("name", "") for k, f in issue_fields.items()}

    def status_at(key, eod):
        """Return the ticket's status as of end-of-day `eod`, reconstructed from
        the changelog. Falls back to current status if there are no transitions.
        Using status-as-of-day (not 'ever reached a milestone') means a ticket
        kicked back from Testing to Code Review correctly re-counts as review
        work — so buildups in Code Review / Testing show on the burndown."""
        trans = by_key.get(key, [])
        if not trans:
            return current_status.get(key, "")
        st = trans[0]["from_status"] or current_status.get(key, "")
        for t in trans:
            if t["timestamp"] and t["timestamp"] <= eod:
                st = t["to_status"]
            else:
                break
        return st

    # ── Committed scope: tickets present within 12-hour grace window ──────────
    grace_cutoff = sprint_start + datetime.timedelta(hours=12)
    first_add_ts = {}
    for sc in scope_changes_all:
        if sc["change"] == "Added" and sc["key"] not in first_add_ts:
            first_add_ts[sc["key"]] = sc["timestamp"]

    initial_committed = {}  # key → sp
    for iss in issues:
        key = iss["key"]
        add_ts = first_add_ts.get(key) or parse_dt(iss["fields"].get("created"))
        if add_ts and add_ts <= grace_cutoff:
            initial_committed[key] = issue_sp.get(key, 0)

    initial_scope   = sum(initial_committed.values())
    committed_count = len(initial_committed)

    # All keys ever in this sprint (grace-window + mid-sprint adds)
    all_sprint_keys = set(initial_committed.keys())
    for sc in scope_changes_all:
        if sc["change"] == "Added":
            all_sprint_keys.add(sc["key"])

    # ── First Done/Closed transition per key (no resolutiondate gate) ─────────
    done_ts = {}
    for t in transitions_all:
        if t["to_status"] in DONE and t["key"] not in done_ts and t["timestamp"]:
            done_ts[t["key"]] = t["timestamp"]

    # ── First Not Required transition per key → treated as scope removal ──────
    nr_ts = {}
    for t in transitions_all:
        if t["to_status"] == "Not Required" and t["key"] not in nr_ts and t["timestamp"]:
            nr_ts[t["key"]] = t["timestamp"]

    # ── Scope delta events bucketed to business days ──────────────────────────
    def snap_to_biz_day(ts):
        """Return Sydney date, rolling weekends forward to Monday."""
        d = ts.astimezone(SYDNEY_OFFSET).date()
        while d.weekday() >= 5:
            d += datetime.timedelta(days=1)
        return d

    scope_delta_by_day = {}

    # Sprint membership changes (after grace window, from scope_changes_all)
    for sc in scope_changes_all:
        ts = sc.get("timestamp_sydney") or sc.get("timestamp")
        if not ts:
            continue
        sp    = issue_sp.get(sc["key"], 0)
        delta = sp if sc["change"] == "Added" else -sp
        d = snap_to_biz_day(ts)
        scope_delta_by_day[d] = scope_delta_by_day.get(d, 0) + delta

    # Not Required transitions → negative scope change
    for key, ts in nr_ts.items():
        if key not in all_sprint_keys:
            continue
        sp = issue_sp.get(key, 0)
        d  = snap_to_biz_day(ts)
        if start_date <= d <= end_date:
            scope_delta_by_day[d] = scope_delta_by_day.get(d, 0) - sp

    # ── Removed-from-sprint timestamps (for dev-complete exclusion) ─────────────
    # Mid-sprint removes subtract SP from running_scope, so they must not also
    # inflate dev_sp or the two cancel and dev_rem stays wrong.
    removed_ts_by_key = {}
    for sc in scope_changes_all:
        if sc["change"] == "Removed":
            ts = sc.get("timestamp") or sc.get("timestamp_sydney")
            if ts and sc["key"] not in removed_ts_by_key:
                removed_ts_by_key[sc["key"]] = ts

    # ── Rework count per day ──────────────────────────────────────────────────
    rework_by_day = {}
    for r in rework_all:
        ts = r.get("timestamp_sydney")
        if ts:
            d = ts.date()
            rework_by_day[d] = rework_by_day.get(d, 0) + 1

    # ── Per-day series ────────────────────────────────────────────────────────
    ideal, dev_rem, review_rem, done_rem, scope_ch, rework_ct = [], [], [], [], [], []
    running_scope = initial_scope

    for i, day in enumerate(biz_days):
        # Apply scope delta for this day before computing remaining
        daily_delta = scope_delta_by_day.get(day, 0)
        running_scope += daily_delta

        ideal_val = initial_scope * (n - 1 - i) / (n - 1) if n > 1 else 0
        ideal.append(round(ideal_val, 1))
        scope_ch.append(daily_delta)
        rework_ct.append(rework_by_day.get(day, 0))

        if day > effective_end:
            dev_rem.append(None)
            review_rem.append(None)
            done_rem.append(None)
            continue

        eod = datetime.datetime.combine(day, datetime.time(23, 59, 59), tzinfo=SYDNEY_OFFSET)

        # Done SP: all sprint tickets with a Done/Closed transition by eod.
        # Tickets set to Not Required by eod are excluded — their scope was removed.
        done_sp = 0
        for key in all_sprint_keys:
            sp = issue_sp.get(key, 0)
            if sp == 0:
                continue
            if key in nr_ts and nr_ts[key] <= eod:
                continue  # scope already removed; don't count as done
            ts = done_ts.get(key)
            if ts and ts <= eod:
                done_sp += sp
        done_rem.append(running_scope - done_sp)

        # Dev-complete & review-complete remaining, based on each ticket's STATUS
        # AS OF eod (not "ever reached") so rework that bounces a ticket back from
        # Testing → Code Review re-counts it as review work, surfacing buildups.
        #   dev_rem    = SP still in BACKLOG ∪ DEV   (status not yet past dev)
        #   review_rem = dev_rem + SP in Code Review (status not yet in Testing)
        # Same NR/removed exclusions as done_sp so removed scope isn't double-counted.
        dev_sp = 0
        review_sp = 0
        for key in all_sprint_keys:
            sp = issue_sp.get(key, 0)
            if sp == 0:
                continue
            if key in nr_ts and nr_ts[key] <= eod:
                continue  # scope already removed; don't count
            rem_ts = removed_ts_by_key.get(key)
            if rem_ts and rem_ts <= eod:
                continue  # removed from sprint; scope already subtracted
            st = status_at(key, eod)
            if st in QA or st in DONE:          # past dev (Code Review / Testing / Done)
                dev_sp += sp
            if st in TEST or st in DONE:        # past code review (Testing / Done)
                review_sp += sp
        dev_rem.append(running_scope - dev_sp)
        review_rem.append(running_scope - review_sp)

    scope_added   = sum(v for v in scope_ch if v > 0)
    scope_removed = abs(sum(v for v in scope_ch if v < 0))
    last_dev    = next((v for v in reversed(dev_rem)    if v is not None), initial_scope)
    last_review = next((v for v in reversed(review_rem) if v is not None), initial_scope)
    last_done   = next((v for v in reversed(done_rem)   if v is not None), initial_scope)
    final_scope = running_scope  # scope at last computed day

    return {
        "biz_days":        biz_days,
        "committed_scope": initial_scope,
        "committed_count": committed_count,
        "final_scope":     final_scope,
        "ideal":      ideal,
        "dev_rem":    dev_rem,
        "review_rem": review_rem,
        "done_rem":   done_rem,
        "scope_ch":   scope_ch,
        "rework":     rework_ct,
        "is_active":     sprint.get("state") == "active",
        "last_dev":      last_dev,
        "last_review":   last_review,
        "last_done":     last_done,
        "scope_added":   scope_added,
        "scope_removed": scope_removed,
        "rework_total":   len(rework_all),
        "rework_tickets": len({r["key"] for r in rework_all}),
        "issue_sp":       issue_sp,
    }


def add_dashboard_sheet(wb, sprint, cs, commitment=None,
                        scope_changes_all=None, transitions_all=None, rework_all=None,
                        cycle_times=None):
    """Add a Dashboard sheet (summary cards + 3 charts + events table) and a hidden Chart Data sheet."""
    if cs is None:
        return

    n        = len(cs["biz_days"])
    spr_name = sprint.get("name", "")

    sprint_start = parse_dt(sprint.get("startDate"))
    sprint_end   = parse_dt(sprint.get("endDate"))
    start_str = sprint_start.astimezone(SYDNEY_OFFSET).strftime("%d %b") if sprint_start else ""
    end_str   = sprint_end.astimezone(SYDNEY_OFFSET).strftime("%d %b %Y") if sprint_end else ""
    date_range = f"  ·  {start_str} – {end_str}" if start_str else ""

    # ── Chart Data sheet (hidden) ─────────────────────────────────────────────
    wsd = wb.create_sheet("Chart Data")
    wsd.append(["Day", "Ideal", "Dev Remaining", "Review Remaining", "Remaining", "Scope Δ", "Rework"])
    for i, day in enumerate(cs["biz_days"]):
        wsd.append([
            day.strftime("%d %b"),
            cs["ideal"][i],
            cs["dev_rem"][i],
            cs["review_rem"][i],
            cs["done_rem"][i],
            cs["scope_ch"][i],
            cs["rework"][i],
        ])
    style_header_row(wsd, max_col=7)
    for col_idx in range(2, 8):
        wsd.column_dimensions[get_column_letter(col_idx)].width = 18
    wsd.column_dimensions["A"].width = 10
    wsd.sheet_state = "hidden"

    # Shared category reference (sprint day labels)
    cats = Reference(wsd, min_col=1, min_row=2, max_row=n + 1)

    # ── Dashboard sheet ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    # Wider columns for cycle time table (cols I–O = 9–15)
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    for _ci in range(12, 16):
        ws.column_dimensions[get_column_letter(_ci)].width = 14

    # Title bar
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = f"Sprint Dashboard  ·  {spr_name}{date_range}"
    title.font  = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    title.fill  = HDR_FILL
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Summary cards (rows 3–5) ──────────────────────────────────────────────
    net_scope   = cs["scope_added"] - cs["scope_removed"]
    final_scope = cs.get("final_scope", cs["committed_scope"])
    sp_closed   = final_scope - int(cs["last_done"])
    sp_dev_done = final_scope - int(cs["last_dev"])

    # Committed card sub-line: show capacity and net scope change if any
    committed_sub_parts = [f"{cs['committed_count']} tickets"]
    if commitment:
        committed_sub_parts.append(f"{commitment['capacity_sp']} SP capacity")
    if net_scope != 0:
        change_str = f"+{net_scope}" if net_scope > 0 else str(net_scope)
        committed_sub_parts.append(f"{change_str} SP scope change")
    committed_sub = "  ·  ".join(committed_sub_parts)

    cards = [
        ("Committed",
         f"{cs['committed_scope']} SP",
         committed_sub,
         "B4B2A9"),
        ("To dev-complete",
         str(int(cs["last_dev"])),
         f"{sp_dev_done} SP reached test",
         "006CA7"),
        ("To done",
         str(int(cs["last_done"])),
         f"{sp_closed} SP closed",
         "6FB544" if cs["is_active"] else "CD442C"),
        ("Scope change",
         f"+{net_scope}" if net_scope > 0 else str(net_scope),
         f"+{cs['scope_added']} added  −{cs['scope_removed']} removed",
         "EAAB30" if net_scope >= 0 else "B4B2A9"),
        ("Test failures",
         str(cs["rework_total"]),
         f"across {cs['rework_tickets']} ticket(s)",
         "E36B5A"),
    ]

    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 14

    for card_i, (label, val, sub, color) in enumerate(cards):
        col_s = card_i * 2 + 1   # 1, 3, 5, 7, 9
        col_e = col_s + 1
        sl, el = get_column_letter(col_s), get_column_letter(col_e)

        ws.merge_cells(f"{sl}3:{el}3")
        lc = ws.cell(3, col_s, label)
        lc.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
        lc.fill = PatternFill("solid", start_color=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{sl}4:{el}4")
        vc = ws.cell(4, col_s, val)
        vc.font   = Font(bold=True, size=16, color="1F2422", name="Arial")
        vc.fill   = PatternFill("solid", start_color="FFFFFF")
        vc.border = Border(left=Side(style="medium", color=color))
        vc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{sl}5:{el}5")
        sc_cell = ws.cell(5, col_s, sub)
        sc_cell.font   = Font(size=8, color="4D504E", name="Arial")
        sc_cell.fill   = PatternFill("solid", start_color="F5F5F5")
        sc_cell.border = Border(left=Side(style="medium", color=color))
        sc_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Burndown line chart ───────────────────────────────────────────────────
    burn = LineChart()
    burn.title          = "Sprint Burndown"
    burn.y_axis.title   = "SP Remaining"
    burn.y_axis.numFmt  = "0"
    burn.height         = 13
    burn.width          = 26
    burn.grouping       = "standard"

    burn.add_data(Reference(wsd, min_col=2, max_col=5, min_row=1, max_row=n + 1),
                  titles_from_data=True)
    burn.set_categories(cats)

    s_ideal = burn.series[0]
    s_ideal.graphicalProperties.line.solidFill = "4D504E"
    s_ideal.graphicalProperties.line.dashStyle = "dash"
    s_ideal.graphicalProperties.line.width = 18000
    s_ideal.marker.symbol = "none"

    s_dev = burn.series[1]
    s_dev.graphicalProperties.line.solidFill = "006CA7"
    s_dev.graphicalProperties.line.width = 22000
    s_dev.marker.symbol = "circle"
    s_dev.marker.size   = 4

    s_review = burn.series[2]
    s_review.graphicalProperties.line.solidFill = "ED7D31"
    s_review.graphicalProperties.line.width = 22000
    s_review.marker.symbol = "triangle"
    s_review.marker.size   = 5

    s_done = burn.series[3]
    s_done.graphicalProperties.line.solidFill = "6FB544"
    s_done.graphicalProperties.line.width = 24000
    s_done.marker.symbol = "square"
    s_done.marker.size   = 4

    # ── Scope-change bar chart (above burndown) ───────────────────────────────
    scope_chart = BarChart()
    scope_chart.title        = "Scope Changes (SP)"
    scope_chart.y_axis.title = "Scope Δ"
    scope_chart.y_axis.numFmt = "0"
    scope_chart.height = 6
    scope_chart.width  = 26
    scope_chart.type   = "col"

    scope_chart.add_data(Reference(wsd, min_col=6, max_col=6, min_row=1, max_row=n + 1),
                         titles_from_data=True)
    scope_chart.set_categories(cats)
    scope_chart.series[0].graphicalProperties.solidFill = "EAAB30"

    ws.add_chart(scope_chart, "A7")

    # ── Burndown line chart (below scope changes) ─────────────────────────────
    ws.add_chart(burn, "A21")

    # ── Rework bar chart ──────────────────────────────────────────────────────
    rework_chart = BarChart()
    rework_chart.title        = "Rework Events (Testing → not Done)"
    rework_chart.y_axis.title = "Count"
    rework_chart.y_axis.numFmt = "0"
    rework_chart.height = 6
    rework_chart.width  = 26
    rework_chart.type   = "col"

    rework_chart.add_data(Reference(wsd, min_col=7, max_col=7, min_row=1, max_row=n + 1),
                          titles_from_data=True)
    rework_chart.set_categories(cats)
    rework_chart.series[0].graphicalProperties.solidFill = "E36B5A"

    ws.add_chart(rework_chart, "A49")

    # ── Sprint Change Events + Sprint Rework Events (side by side) ───────────
    # Layout: Change table cols A-C, gap col D, Rework table cols E-G.
    # Rework chart at A49, height=6cm ≈ 11 rows → tables start at row 63.
    issue_sp = cs.get("issue_sp", {})

    def _fmt_event_dt(ts):
        if ts is None:
            return ""
        if hasattr(ts, "astimezone"):
            ts = ts.astimezone(SYDNEY_OFFSET)
        return ts.strftime("%d %b %Y %H:%M")

    # ── Build change rows ─────────────────────────────────────────────────────
    change_rows = []
    for sc in (scope_changes_all or []):
        sp = issue_sp.get(sc["key"], 0)
        change_rows.append({
            "key":  sc["key"],
            "date": _fmt_event_dt(sc.get("timestamp_sydney") or sc.get("timestamp")),
            "sp":   sp if sc["change"] == "Added" else -sp,
        })
    for t in (transitions_all or []):
        if t["to_status"] == "Not Required":
            change_rows.append({
                "key":  t["key"],
                "date": _fmt_event_dt(t.get("timestamp_sydney")),
                "sp":   -(issue_sp.get(t["key"], 0)),
            })
    change_rows.sort(key=lambda x: x["key"])

    # ── Build rework rows ─────────────────────────────────────────────────────
    rework_rows = []
    for r in (rework_all or []):
        rework_rows.append({
            "key":  r["key"],
            "date": _fmt_event_dt(r.get("timestamp_sydney")),
            "sp":   issue_sp.get(r["key"], 0),
        })
    rework_rows.sort(key=lambda x: x["key"])

    # ── Table header colours ──────────────────────────────────────────────────
    CHANGE_HDR_FILL = PatternFill("solid", start_color="FFC000")   # amber yellow
    REWORK_HDR_FILL = PatternFill("solid", start_color="C00000")   # solid red
    CHANGE_HDR_FONT = Font(bold=True, color="000000", name="Arial", size=10)
    REWORK_HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def _write_side_table(title, rows, start_col, hdr_fill, hdr_font, row_highlight):
        """Write a 3-column (Issue ID, Date, SP) table starting at tbl_start_row."""
        r = tbl_start
        sl = get_column_letter(start_col)
        el = get_column_letter(start_col + 2)

        ws.merge_cells(f"{sl}{r}:{el}{r}")
        tc = ws.cell(r, start_col, title)
        tc.font = hdr_font
        tc.fill = hdr_fill
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        r += 1

        for ci_off, h in enumerate(["Issue ID", "Date (AEST)", "SP"]):
            hc = ws.cell(r, start_col + ci_off, h)
            hc.fill = hdr_fill
            hc.font = hdr_font
            hc.border = THIN_BORDER
            hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

        if rows:
            for i, ev in enumerate(rows):
                ws.cell(r, start_col).value     = ev["key"]
                ws.cell(r, start_col + 1).value = ev["date"]
                ws.cell(r, start_col + 2).value = ev["sp"]
                for ci_off in range(3):
                    c = ws.cell(r, start_col + ci_off)
                    c.font   = BODY_FONT
                    c.border = THIN_BORDER
                    c.alignment = Alignment(
                        horizontal="left" if ci_off == 0 else "center",
                        vertical="center",
                    )
                    if i % 2 == 1:
                        c.fill = row_highlight
                r += 1
        else:
            c = ws.cell(r, start_col, "None")
            c.font = Font(italic=True, color="999999", name="Arial", size=9)

    def _write_cycle_time_table(ct_data, start_col):
        """Write a 5-column cycle time table: Issue ID, In Progress, Blocked, Code Review, Test."""
        CYCLE_HDR_FILL = PatternFill("solid", start_color="1F4E79")
        CYCLE_HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        col_headers = ["Issue ID", "SP Estimate", "Total Cycle Time SP", "In Progress", "Blocked", "Code Review", "Test"]
        n_cols = len(col_headers)
        sl = get_column_letter(start_col)
        el = get_column_letter(start_col + n_cols - 1)

        r = tbl_start
        ws.merge_cells(f"{sl}{r}:{el}{r}")
        tc = ws.cell(r, start_col, "Cycle Time Per Item - hours (Story Points)")
        tc.font      = CYCLE_HDR_FONT
        tc.fill      = CYCLE_HDR_FILL
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        r += 1

        for ci_off, h in enumerate(col_headers):
            hc = ws.cell(r, start_col + ci_off, h)
            hc.fill      = CYCLE_HDR_FILL
            hc.font      = CYCLE_HDR_FONT
            hc.border    = THIN_BORDER
            hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

        def _fmt_ct(hours):
            """Format as '12.5 (3)' — business hours + nearest-Fibonacci SP estimate.
            2 SP/day × 8 h/day → SP = hours / 4.  Returns None when hours is zero."""
            if not hours:
                return None
            sp = _nearest_fibonacci(hours / 4)
            return f"{hours:.1f} ({sp})"

        sorted_keys = sorted(ct_data.keys())
        if sorted_keys:
            for i, key in enumerate(sorted_keys):
                totals = ct_data[key]
                actual_sp = sum(
                    _nearest_fibonacci(totals[b] / 4)
                    for b in ("In Progress", "Code Review", "Test")
                    if totals[b]
                ) or None
                vals = [
                    key,
                    issue_sp.get(key) or None,
                    actual_sp,
                    _fmt_ct(totals["In Progress"]),
                    _fmt_ct(totals["Blocked"]),
                    _fmt_ct(totals["Code Review"]),
                    _fmt_ct(totals["Test"]),
                ]
                sp_est = vals[1]   # SP Estimate (numeric)
                act_sp = vals[2]   # Actual SP   (numeric)
                ACTUAL_WARN_FILL = PatternFill("solid", start_color="FFF2CC")  # pale yellow
                ACTUAL_OVER_FILL = PatternFill("solid", start_color="FFCCCC")  # pale red

                for ci_off, v in enumerate(vals):
                    c = ws.cell(r, start_col + ci_off, v)
                    c.font   = BODY_FONT
                    c.border = THIN_BORDER
                    c.alignment = Alignment(
                        horizontal="left" if ci_off == 0 else "center",
                        vertical="center",
                    )
                    if ci_off == 2 and act_sp and sp_est:
                        if act_sp > sp_est * 1.3:
                            c.fill = ACTUAL_OVER_FILL
                        elif act_sp > sp_est:
                            c.fill = ACTUAL_WARN_FILL
                        elif i % 2 == 1:
                            c.fill = ALT_FILL
                    elif i % 2 == 1:
                        c.fill = ALT_FILL
                r += 1
        else:
            c = ws.cell(r, start_col, "No data")
            c.font = Font(italic=True, color="999999", name="Arial", size=9)

    note_row = 63
    note_cell = ws.cell(note_row, 1,
        "Events in the below tables may have occured outside the bounds of the current sprint "
        "and will not be reflected on the charts above. Please review the date an event occured.")
    note_cell.font = Font(italic=True, color="555555", name="Arial", size=9)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=13)
    ws.row_dimensions[note_row].height = 28

    tbl_start = 65
    _write_side_table("Sprint Change Events", change_rows,
                      start_col=1, hdr_fill=CHANGE_HDR_FILL, hdr_font=CHANGE_HDR_FONT,
                      row_highlight=WARN_FILL)
    _write_side_table("Sprint Rework Events", rework_rows,
                      start_col=5, hdr_fill=REWORK_HDR_FILL, hdr_font=REWORK_HDR_FONT,
                      row_highlight=REWORK_FILL)
    if cycle_times:
        _write_cycle_time_table(cycle_times, start_col=9)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Extract Jira sprint data to Excel + chat markdown.")
    parser.add_argument("--project", "-p", required=True, help="Jira project key, e.g. ADA")
    parser.add_argument("--sprint",  "-s", required=True, help="'active', sprint name substring, or numeric sprint ID")
    parser.add_argument("--base-url", default=DEFAULTS["base_url"], help="Jira base URL")
    parser.add_argument("--email",   default=DEFAULTS["email"],    help="Atlassian account email")
    parser.add_argument("--token",   default=DEFAULTS["token"],    help="Jira API token")
    parser.add_argument("--out-dir", default=".", help="Output directory")
    parser.add_argument("--commitment-page", "-c", default=None,
                        help="Confluence commitment page title (auto-matched from sprint name if omitted)")
    args = parser.parse_args()

    if not args.email or not args.token:
        sys.exit(
            "ERROR: Jira credentials not set.\n"
            "Either pass --email and --token, or set JIRA_EMAIL and JIRA_API_TOKEN environment variables.\n"
            "Get an API token at: https://id.atlassian.com/manage-profile/security/api-tokens"
        )

    session = make_session(args.email, args.token)

    print(f"Resolving sprint '{args.sprint}' for project {args.project}...")
    sprint = resolve_sprint(session, args.base_url, args.project, args.sprint)
    print(f"  Found: {sprint['name']} (ID {sprint['id']}, {sprint['state']})")

    print("Fetching issues with changelog...")
    issues = fetch_issues_with_changelog(session, args.base_url, args.project, sprint["name"])
    print(f"  {len(issues)} issues fetched.")

    print("Extracting transitions and scope changes...")
    transitions_all = []
    rework_all = []
    scope_changes_all = []

    sprint_start = parse_dt(sprint.get("startDate"))
    grace_cutoff = sprint_start + datetime.timedelta(hours=12) if sprint_start else None

    for issue in issues:
        transitions = extract_transitions(issue)
        transitions_all.extend(transitions)

        for t in transitions:
            if t["from_status"] == "Testing" and t["to_status"] != "Done":
                rework_all.append(t)

        scope_changes = extract_sprint_membership_changes(issue, sprint["id"])
        for sc in scope_changes:
            # Only count as mid-sprint if it happened after the grace window
            if grace_cutoff and sc["timestamp"] and sc["timestamp"] > grace_cutoff:
                scope_changes_all.append(sc)

    transitions_all.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))
    rework_all.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))
    scope_changes_all.sort(key=lambda x: x["timestamp"] or datetime.datetime.min.replace(tzinfo=datetime.timezone.utc))

    print(f"  {len(transitions_all)} transitions, {len(rework_all)} rework events, {len(scope_changes_all)} scope changes.")

    print("Fetching sprint commitment from Confluence...")
    commitment = fetch_sprint_commitment(session, args.base_url, args.project, sprint["name"],
                                         args.commitment_page)
    if commitment:
        print(f"  {len(commitment['members'])} members — "
              f"{commitment['planned_excl_ncc']} counted days → {commitment['capacity_sp']} SP capacity")
    else:
        print("  No commitment data found (Confluence page not matched or table not parsed).")

    # Build slug for filenames
    slug = re.sub(r"[^A-Za-z0-9_\-]", "_", sprint["name"])[:40]
    today = datetime.date.today().isoformat()
    base_name = f"{args.project}_{slug}_{today}"
    out_dir = args.out_dir

    # Write Excel
    print("Building Excel workbook...")
    wb = build_workbook(sprint, issues, transitions_all, rework_all, scope_changes_all, commitment)
    xlsx_path = os.path.join(out_dir, base_name + ".xlsx")
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

    # Write chat markdown
    print("Building chat summary markdown...")
    md = build_chat_summary(sprint, issues, transitions_all, rework_all, scope_changes_all, commitment)
    md_path = os.path.join(out_dir, base_name + "_chat.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"  Saved: {md_path}")

    print("\nDone.")
    print(f"  Excel  → {xlsx_path}")
    print(f"  Chat   → {md_path}  (paste this into the burndown conversation)")
    print(f"\n  Rework events: {len(rework_all)}")
    if rework_all:
        from collections import Counter
        counts = Counter(r["key"] for r in rework_all)
        for key, n in counts.most_common():
            print(f"    {key}: {n} failure{'s' if n>1 else ''}")


if __name__ == "__main__":
    main()
